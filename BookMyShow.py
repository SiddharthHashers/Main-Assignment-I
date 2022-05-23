import openpyxl
import xlsxwriter
import xlwt
import xlrd
from xlrd import sheet
import pandas as pd
from xlrd.timemachine import xrange
from xlutils.copy import copy
from xlwt import Workbook

import RegisterNewAccount


class BookmyShow():
    def __init__(self):
        self.uservale=0
        print('******Welcome to BookMyShow******* ')
        print('1) Login')
        print('2) Register new account')
        print('3) Exit')
        n = int(input('Enter a you action:'))
        admin=1
        if n==admin :
            n4 = input('You are Admin or User :')
            if n4 == 'admin':
                user = input('Enter Username:  ')
                password = input('Enter password:  ')
                if user == 'admin' and password == 'admin@123':
                    print('******Welcome Admin*******')
                    print("1) Add New movie info ")
                    print('2) Edit Movie Info ')
                    print('3) Delete Movies ')
                    print('4) Logout enter ')
                    num = int(input("Enter a you action: "))
                    global p
                    p=0
                    if num == 1:
                        wb = Workbook()
                        # add_sheet is used to create sheet.
                        sheet1 = wb.add_sheet('Sheet 1')
                        print("Number of New Movies Need To Be Entered")
                        size=int(input())
                        self.uservale=size
                        for i in range(1,size+1):
                            p=1
                            print("Enter Title")
                            Title=str(input())
                            if i == 1:
                                sheet1.write(0, 0, 'Title')
                                sheet1.write(0, i, Title)
                            else:
                                sheet1.write(0, i, Title)
                            print("Enter Genre")
                            Genre=str(input())
                            if i == 1:
                                sheet1.write(1, 0, 'Genre')
                                sheet1.write(1, i, Genre)
                            else:
                                sheet1.write(1, i, Genre)

                            print("enter Length")
                            Length=str(input())
                            if i == 1:
                                sheet1.write(2, 0, 'Length')
                                sheet1.write(2, i, Length)
                            else:
                                sheet1.write(2, i, Length)
                            print("enter cast")
                            Cast=str(input())
                            if i == 1:
                                sheet1.write(3, 0, 'Cast')
                                sheet1.write(3, i, Cast)
                            else:
                                sheet1.write(3, i, Cast)
                            print("enter director")
                            Director=str(input())
                            if i == 1:
                                sheet1.write(4, 0, 'Director')
                                sheet1.write(4, i, Director)
                            else:
                                sheet1.write(4, i, Director)
                            print("Admin rating ")
                            AdminRating=str(input())
                            if i == 1:
                                sheet1.write(5, 0, 'AdminRating')
                                sheet1.write(5, i, AdminRating)
                            else:
                                sheet1.write(5, i, AdminRating)
                            print("Language")
                            Language=str(input())
                            if i == 1:
                                sheet1.write(6, 0, 'Language')
                                sheet1.write(6, i, Language)
                            else:
                                sheet1.write(6, i, Language)
                            print("timings")
                            Timings=str(input())
                            if i == 1:
                                sheet1.write(7, 0, 'Timings')
                                sheet1.write(7, i, Timings)
                            else:
                                sheet1.write(7, i, Timings)
                            print("number of Shows in a day")
                            NumberofShowsinaday=str(input())
                            if i == 1:
                                sheet1.write(8, 0, 'NumberofShowsinaday')
                                sheet1.write(8, i, NumberofShowsinaday)
                            else:
                                sheet1.write(8, i, NumberofShowsinaday)
                            print("first show at")
                            FirstShow=str(input())
                            if i == 1:
                                sheet1.write(9, 0, 'FirstShow')
                                sheet1.write(9, i, FirstShow)
                            else:
                                sheet1.write(9, i, FirstShow)
                            print("intervaltime")
                            IntervalTime=str(input())
                            if i == 1:
                                sheet1.write(10, 0, 'IntervalTime')
                                sheet1.write(10, i, IntervalTime)
                            else:
                                sheet1.write(10, i, IntervalTime)
                            print("Gap betweeen shows")
                            GapBetweenShows=str(input())
                            if i == 1:
                                sheet1.write(11, 0, 'GapBetweenShows')
                                sheet1.write(11, i, GapBetweenShows)
                            else:
                                sheet1.write(11, i, GapBetweenShows)
                            print("capacity")
                            Capacity=str(input())
                            if i == 1:
                                sheet1.write(12, 0, 'Capacity')
                                sheet1.write(12, i, Capacity)
                            else:
                                sheet1.write(12, i, Capacity)
                            wb.save('MOVIEDETAILS.xls')
                        if p == 1:
                            loc = ("/Users/sikalidas/PycharmProjects/20May/MOVIEDETAILS.xls")
                            wb = xlrd.open_workbook(loc)
                            sheet = wb.sheet_by_index(0)
                            s = sheet.cell_value(0, 1)
                            rows = sheet.nrows
                            collumns = sheet.ncols
                            # below is the conversion from xls to xlsx
                            loc = ("/Users/sikalidas/PycharmProjects/20May/tempadmin.xlsx")
                            wb = openpyxl.load_workbook("tempadmin.xlsx")
                            login_sheet = wb['Sheet1']
                            ls_max_row1 = login_sheet.max_row  # max row in login sheet
                            ls_max_col1 = login_sheet.max_column
                            for i in range(0, rows):
                                for j in range(collumns):
                                    copyvalue = sheet.cell_value(i, j)
                                    login_sheet.cell(row=i + 1, column=j + 1, value=copyvalue)
                                    wb.save("tempadmin.xlsx")
                    elif num == 2:
                        print('******Welcome Admin******* ')
                        print('Select movie which you want to edit:')
                        print("**Below are the list of movie please enter the name case sensitive")
                        loc = ("/Users/sikalidas/PycharmProjects/20May/tempadmin.xlsx")
                        wb = openpyxl.load_workbook("tempadmin.xlsx")
                        login_sheet = wb['Sheet1']
                        ls_max_row1 = login_sheet.max_row  # max row in login sheet
                        ls_max_col1 = login_sheet.max_column
                        for i in range(1, ls_max_col1 + 1):
                            sheetvalue = login_sheet.cell(1, i).value
                            print(sheetvalue)
                        user = input('Enter Movie name:  ')
                        global selectedcol
                        loc = ("/Users/sikalidas/PycharmProjects/20May/tempadmin.xlsx")
                        wb1 = openpyxl.load_workbook("tempadmin.xlsx")
                        login_sheet1 = wb1['Sheet1']
                        ls_max_row1 = login_sheet1.max_row  # max row in login sheet
                        ls_max_col1 = login_sheet1.max_column
                        for i in range(1, ls_max_col1+1):
                            sheetvalue = login_sheet1.cell(1, i).value
                            if user == sheetvalue:
                                selectedcol=i
                        global o
                        o =1
                        for i in range(1,ls_max_row1+1):
                            save=login_sheet1.cell(i, 1).value
                            print("Enter Changes For "+ str(save))
                            Title = str(input())
                            login_sheet1.cell(row=i, column=selectedcol, value=Title)
                            wb1.save("tempadmin.xlsx")
                    elif num == 3:
                        print('******Welcome Admin******* ')
                        print('Title of the movie to be deleted: ')
                        loc = ("/Users/sikalidas/PycharmProjects/20May/tempadmin.xlsx")
                        wb1 = openpyxl.load_workbook("tempadmin.xlsx")
                        login_sheet1 = wb1['Sheet1']
                        ls_max_row1 = login_sheet1.max_row  # max row in login sheet
                        ls_max_col1 = login_sheet1.max_column
                        print("**Below are the list of movie please enter the name case sensitive")
                        for i in range(1, ls_max_col1 + 1):
                            sheetvalue = login_sheet1.cell(1, i).value
                            print(sheetvalue)
                        global refname
                        refname=0
                        user = input('Enter Movie name:  ')
                        for i in range(1, ls_max_col1 + 1):
                            sheetvalue = login_sheet1.cell(1, i).value
                            if user == sheetvalue:
                                refname=i
                        for i in range(1, ls_max_row1+1):
                            login_sheet1.cell(row=i, column=refname, value='')
                            wb1.save("tempadmin.xlsx")
                        print("Movie is Deleted : ")
                    elif num == 4:
                        print("You Have Successfully Loged Out Thank You : ")
                else:
                    print()
                    print("Please Retry with Correct Username or Password: ")
            elif n4 == 'user':
                cd1 = RegisterNewAccount.Userprofile()
                cd1.userlogin()
        elif n==2:
            cd=RegisterNewAccount.Userprofile()
            cd.userdetails()
        elif n==3:
            print("exit")
        else:
            print("Please enter correct choice ")
obj_a=BookmyShow()
print()