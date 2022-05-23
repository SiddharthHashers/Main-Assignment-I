import openpyxl
import xlrd

import UserActions


class Userprofile():
    def __init__(self):
        self.flag = None
        self.verifyvalue=0
        self.tempt=0

    def userlogin(self):

        print("******Welcome to BooyMyShow******* ")
        loc = ("/Users/sikalidas/PycharmProjects/20May/RegUserData.xlsx")
        wb = openpyxl.load_workbook("RegUserData.xlsx")
        login_sheet = wb['Sheet1']
        ls_max_row = login_sheet.max_row  # max row in login sheet
        ls_max_col = login_sheet.max_column
        user = input('Enter Username:  ')
        password = input("Enter Password: ")
        self.flag = 0
        for i in range(1, ls_max_row+1):
            global var
            global username
            var1 = login_sheet.cell(i, 1).value
            var2 = login_sheet.cell(i, 2).value
            var = var1
            if var == user and var2 == password:
                loc = ("/Users/sikalidas/PycharmProjects/20May/MOVIEDETAILS.xls")
                wb = xlrd.open_workbook(loc)
                sheet = wb.sheet_by_index(0)
                s = sheet.cell_value(0, 1)
                rows = sheet.nrows
                collumns = sheet.ncols
                self.flag = 1;
                username = var1
        if self.flag == 1:
            print("******Welcome" + ' ' + str(username) + "******* ")
            print('Below Are The List Of Movies Listed. ')
            loc = ("/Users/sikalidas/PycharmProjects/20May/tempadmin.xlsx")
            wb = openpyxl.load_workbook("tempadmin.xlsx")
            login_sheet = wb['Sheet1']
            ls_max_row1 = login_sheet.max_row  # max row in login sheet
            ls_max_col1 = login_sheet.max_column
            for i in range(2, ls_max_col1 + 1):
                sheetvalue = login_sheet.cell(1, i).value
                if sheetvalue != None:
                    print(str(i) + ')' + ' ' + sheetvalue)
            print('4) logout')
            user1 = input('Enter the Options Name By Entering Words : ')
            if user1 == 'logout':
                print()
                print('Thank you for Visiting' + ' ' + str(username))
            for c in range(2, ls_max_col1 + 1):
                sheetvalues = login_sheet.cell(1, c).value
                if user1 == sheetvalues:
                    global iint
                    iint = c
                    self.verifyvalue = 1
                    break
            if self.verifyvalue == 1:
                for p in range(1, ls_max_row1+1):
                    sheetvalue1 = login_sheet.cell(p, 1).value
                    sheetvalue = login_sheet.cell(p, iint).value
                    print(sheetvalue1 + ':' + '' + sheetvalue)
                    self.tempt=1
            if self.tempt == 1:
                print('User Rating: 3.5/10')
                print('1) Book Tickets')
                print('2) Cancel Tickets')
                print('3) Give User Rating')
                takevalue=int(input("Enter your Choice :"))
                if takevalue == 1:
                    cd= UserActions
                    cd.bookings(username)
                if takevalue == 2:
                    cd=UserActions
                    cd.canclebooking(username)
                if takevalue == 3:
                    cd= UserActions
                    cd.ratings(username)
        else:
            print()
            print('Please Create New Account To Proceed : ')


    def userdetails(self):
        print("******  ****Create new Account*****  **********")
        loc = ("/Users/sikalidas/PycharmProjects/20May/RegUserData.xlsx")
        wb = openpyxl.load_workbook("RegUserData.xlsx")
        login_sheet = wb['Sheet1']
        creat_user = wb['Sheet2']
        ls_max_row = login_sheet.max_row  # max row in login sheet
        ls_max_col = login_sheet.max_column
        user = input('Enter Username:  ')
        password = input('Enter Password:  ')
        for i in range(1, ls_max_row+1):
            var1 = login_sheet.cell(i, 1).value
            var2 = login_sheet.cell(i, 2).value
            var = var1
            if var == user and var2 == password:
                self.flag = 1
                break
        if self.flag == 1:
            print('User already exists')
        if self.flag != 1:
            loc = ("/Users/sikalidas/PycharmProjects/20May/RegUserData.xlsx")
            wb = openpyxl.load_workbook("RegUserData.xlsx")
            login_sheet = wb['Sheet1']
            login_sheet.cell(row=ls_max_row + 1, column=1, value=user)
            login_sheet.cell(row=ls_max_row + 1, column=2, value=password)
            wb.save("RegUserData.xlsx")
            print("Please Enter Below Additional Details")
            Name = input('Enter Name:  ')
            Email = input('Enter Email:  ')
            Phone = input('Enter Phone:  ')
            Age = input('Enter Age:  ')
            max_row = creat_user.max_row  # max row in login sheet
            max_col = creat_user.max_column
            a2 = 2
            it=1
            for i in range(1, max_row+1):
                if i == 1:
                    creat_user.cell(row=i, column=a2, value=Name)
                    wb.save("RegUserData.xlsx")
                if i == 2:
                    creat_user.cell(row=i, column=a2, value=Email)
                    wb.save("RegUserData.xlsx")
                if i == 3:
                    creat_user.cell(row=i, column=a2, value=Phone)
                    wb.save("RegUserData.xlsx")
                if i == 4:
                    creat_user.cell(row=i, column=a2, value=Age)
                    wb.save("RegUserData.xlsx")

