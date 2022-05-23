import openpyxl
import xlrd


def bookings(name):
    print("******Welcome" + ' ' + str(name) + "******* ")
    loc = ("/Users/sikalidas/PycharmProjects/20May/MOVIEDETAILS.xls")
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    rows = sheet.nrows
    collumns = sheet.ncols
    print('Below Are The List Of Movies Listed. ')
    loc = ("/Users/sikalidas/PycharmProjects/20May/tempadmin.xlsx")
    wb = openpyxl.load_workbook("tempadmin.xlsx")
    login_sheet = wb['Sheet1']
    ls_max_row1 = login_sheet.max_row  # max row in login sheet
    ls_max_col1 = login_sheet.max_column
    for i in range(2, ls_max_col1 + 1):
        sheetvalue = login_sheet.cell(1, i).value
        if sheetvalue != None:
            print(str(i) + ')' + ' ' + sheetvalue)
    global temp
    global collum
    global countofseat
    strvalue1 = input('Select the Movie By Word ')
    for j in range(2, ls_max_col1 + 1):
        sheetvalue = login_sheet.cell(1, j).value
        if sheetvalue == strvalue1:
            temp = 1
            collum = j
    if temp == 1:
        sheetvalue = login_sheet.cell(8, collum).value
        list = sheetvalue.split(",")
        print(list)
        for i in range(len(list)):
            j = str(i + 1)
            print(j + " :" + list[i])
        t = int(input("Select Timings: "))
        print("Timing : " + list[t-1])
        totalseatsheet1 = login_sheet.cell(13, collum).value
        loc = ("/Users/sikalidas/PycharmProjects/20May/RegUserData.xlsx")
        wb = openpyxl.load_workbook("RegUserData.xlsx")
        login_sheet = wb['Sheet3']
        savetemp=int(totalseatsheet1)
        login_sheet.cell(row=1, column=2, value=savetemp)
        wb.save("RegUserData.xlsx")
        totalseatsheet2 = login_sheet.cell(1, 2).value
        print('Remaining Seats: ' + str(totalseatsheet2))
        seaatbooking = int(input("Enter Number of seats: "))
        countofseat =  totalseatsheet2 - seaatbooking
        loc = ("/Users/sikalidas/PycharmProjects/20May/RegUserData.xlsx")
        wb = openpyxl.load_workbook("RegUserData.xlsx")
        login_sheet = wb['Sheet3']
        login_sheet.cell(row=2, column=2, value=seaatbooking)
        login_sheet.cell(row=3, column=2, value=countofseat)
        wb.save("RegUserData.xlsx")
        print("Thanks for booking. ")
        # login_sheet.cell(row=ls_max_row + 1, column=1, value=user)


def canclebooking(name):
    print("******Welcome" + ' ' + str(name) + "******* ")
    loc = ("/Users/sikalidas/PycharmProjects/20May/RegUserData.xlsx")
    wb = openpyxl.load_workbook("RegUserData.xlsx")
    login_sheet = wb['Sheet3']
    countofbookedseat = login_sheet.cell(2, 2).value
    remaingseat=login_sheet.cell(3,2).value
    print('Confirmed Seats : '+ str(countofbookedseat))
    print('Remaining seats: '+ str(remaingseat))
    Cancelseat = int(input("Number of seats you want to cancel:"))
    login_sheet.cell(row=4, column=2, value=Cancelseat)
    wb.save("RegUserData.xlsx")
    coutaftercancel=countofbookedseat-Cancelseat
    addseat=Cancelseat+remaingseat
    login_sheet.cell(row=2, column=2, value=coutaftercancel)
    login_sheet.cell(row=3, column=2, value=addseat)
    wb.save("RegUserData.xlsx")



def ratings(name):
    print("******Welcome" + ' ' + str(name) + "******* ")
    strr=input('Please enter rating for the following movie: ')
    print('Updated Rating for Movie is : '+strr)


class UserEdits():
    def __init__(self):
        pass
